#!/usr/bin/env python3
"""
CSU Steam Bill Automation App

Desktop tool to:
- read a Corix Cleveland Thermal steam bill PDF
- parse meter rows and Current Charges
- update the team's trusted Excel workbook/template
- preserve existing workbook formulas and summary logic
- generate an optional validation report

Important business rules:
- Ignore trailing alphabetic characters in meter reading values (e.g. 393,115E -> 393115)
- Always use the multiplier from the bill
- Use Current Charges as the bill amount
- Update only workbook input cells; do not replace workbook formulas
"""

from __future__ import annotations

import json
import re
import shutil
import sys
import traceback
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

from openpyxl import load_workbook

try:
    import pdfplumber
except Exception as exc:
    raise SystemExit(
        "Missing dependency: pdfplumber. Install it first with:\n"
        "py -m pip install pdfplumber\n"
        f"Original error: {exc}"
    )


CURRENT_CHARGES_RE = re.compile(
    r"Current Charges\s+\$?\s*([\d,]+\.\d{2})",
    re.IGNORECASE,
)

TOTAL_USAGE_RE = re.compile(
    r"Total Usage\s+([\d,]+(?:\.\d+)?)\s+LB",
    re.IGNORECASE,
)

DATE_RANGE_RE = re.compile(
    r"([A-Za-z]+\s+\d{1,2}\s*-\s*[A-Za-z]+\s+\d{1,2},\s+\d{4})"
)

# Corix rows are often extracted with line breaks in the middle.
# So we parse against normalized text, not one line at a time.
#
# Supported shapes:
# 1) meter prev_date prev curr_date curr usage LB multiplier
# 2) meter prev_date prev curr_date curr multiplier usage LB
#
# Example:
# 1421 12/31/25 8,871 01/31/26 9,813 30,615 LB 32.50
# or
# 1421 12/31/25 8,871 01/31/26 9,813 32.50 30,615 LB

METER_INFO_ROW_RE = re.compile(
    r"^(?P<meter>\d{3,10})\s+"
    r"(?P<prev_date>\d{2}/\d{2}/\d{2})\s+"
    r"(?P<prev>[\d,]+(?:\.\d+)?[A-Za-z]*)\s+"
    r"(?P<curr_date>\d{2}/\d{2}/\d{2})\s+"
    r"(?P<curr>[\d,]+(?:\.\d+)?[A-Za-z]*)\s+"
    r"(?P<usage>[\d,]+(?:\.\d+)?)\s+LB\s+"
    r"(?P<multiplier>\d+(?:\.\d+)?)$",
    re.IGNORECASE,
)



@dataclass
class MeterBillRow:
    meter: str
    prev_read: int
    curr_read: int
    multiplier: float
    bill_usage_lbs: float
    prev_date: str
    curr_date: str
    raw_prev: str
    raw_curr: str


@dataclass
class ValidationIssue:
    level: str
    message: str


@dataclass
class GenerationResult:
    output_workbook: str
    report_path: Optional[str]
    matched_meter_count: int
    workbook_meter_count: int
    bill_meter_count: int
    missing_in_bill: List[str]
    extra_in_bill: List[str]
    multiplier_changes: List[Dict[str, object]]
    usage_mismatches: List[Dict[str, object]]
    bill_total_usage_lbs: float
    workbook_total_usage_lbs_from_inputs: float
    current_charges: float
    bill_period: Optional[str]
    issues: List[ValidationIssue]


def clean_numeric_text(value: str) -> str:
    return re.sub(r"[^0-9.]", "", value or "")


def parse_reading(value: str) -> int:
    cleaned = clean_numeric_text(value)
    if not cleaned:
        raise ValueError(f"Could not parse meter reading from {value!r}")
    return int(float(cleaned))


def parse_number(value: str) -> float:
    cleaned = clean_numeric_text(value)
    if not cleaned:
        raise ValueError(f"Could not parse number from {value!r}")
    return float(cleaned)


def normalize_pdf_text(text: str) -> str:
    text = text.replace("\xa0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n+", "\n", text)
    return text.strip()


def read_pdf_text(pdf_path: Path) -> str:
    parts: List[str] = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                parts.append(page_text)
    text = "\n".join(parts)
    if not text.strip():
        raise ValueError(
            "No text could be extracted from the PDF. "
            "The PDF may be scanned or image-based."
        )
    return normalize_pdf_text(text)




def parse_bill(pdf_path: Path) -> Tuple[Dict[str, MeterBillRow], float, float, Optional[str]]:
    text = read_pdf_text(pdf_path)

    current_charges_match = CURRENT_CHARGES_RE.search(text)
    if not current_charges_match:
        raise ValueError("Could not find 'Current Charges' in the PDF.")
    current_charges = parse_number(current_charges_match.group(1))

    total_usage_match = TOTAL_USAGE_RE.search(text)
    if not total_usage_match:
        raise ValueError("Could not find 'Total Usage' in the PDF.")
    total_usage_lbs = parse_number(total_usage_match.group(1))

    period_match = DATE_RANGE_RE.search(text)
    bill_period = period_match.group(1) if period_match else None

    rows: Dict[str, MeterBillRow] = {}

    for raw_line in text.splitlines():
        line = re.sub(r"\s+", " ", raw_line).strip()
        match = METER_INFO_ROW_RE.match(line)
        if not match:
            continue

        meter = match.group("meter")
        prev_read = parse_reading(match.group("prev"))
        curr_read = parse_reading(match.group("curr"))
        multiplier = float(match.group("multiplier"))
        bill_usage_lbs = parse_number(match.group("usage"))

        rows[meter] = MeterBillRow(
            meter=meter,
            prev_read=prev_read,
            curr_read=curr_read,
            multiplier=multiplier,
            bill_usage_lbs=bill_usage_lbs,
            prev_date=match.group("prev_date"),
            curr_date=match.group("curr_date"),
            raw_prev=match.group("prev"),
            raw_curr=match.group("curr"),
        )

    if not rows:
        raise ValueError(
            "No meter rows were parsed from the PDF. "
            "Check the Corix bill format or extracted text."
        )

    return rows, current_charges, total_usage_lbs, bill_period


def normalize_meter_id(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        return str(int(float(text)))
    return None


def find_bill_amount_target(ws) -> Tuple[int, int]:
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if not isinstance(val, str):
                continue
            label = val.strip().lower()
            if "cost on bill" in label or label == "bill amount":
                return cell.row + 1, cell.column
    return 53, 1


def find_meter_rows(ws) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for row_idx in range(1, ws.max_row + 1):
        meter = normalize_meter_id(ws.cell(row_idx, 1).value)
        if meter:
            mapping[meter] = row_idx
    return mapping


def safe_float(value: object) -> Optional[float]:
    if value is None:
        return None
    try:
        return float(value)
    except Exception:
        return None


def generate_validation_report(
    report_path: Optional[Path],
    result: GenerationResult,
    bill_rows: Dict[str, MeterBillRow],
) -> Optional[str]:
    if report_path is None:
        return None

    payload = {
        "summary": {
            "output_workbook": result.output_workbook,
            "matched_meter_count": result.matched_meter_count,
            "workbook_meter_count": result.workbook_meter_count,
            "bill_meter_count": result.bill_meter_count,
            "bill_total_usage_lbs": result.bill_total_usage_lbs,
            "workbook_total_usage_lbs_from_inputs": result.workbook_total_usage_lbs_from_inputs,
            "current_charges": result.current_charges,
            "bill_period": result.bill_period,
        },
        "missing_in_bill": result.missing_in_bill,
        "extra_in_bill": result.extra_in_bill,
        "multiplier_changes": result.multiplier_changes,
        "usage_mismatches": result.usage_mismatches,
        "issues": [asdict(issue) for issue in result.issues],
        "bill_rows": {meter: asdict(row) for meter, row in bill_rows.items()},
    }
    report_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    return str(report_path)


def update_workbook_from_bill(
    pdf_path: Path,
    workbook_path: Path,
    output_path: Path,
    report_path: Optional[Path] = None,
) -> GenerationResult:
    bill_rows, current_charges, bill_total_usage_lbs, bill_period = parse_bill(pdf_path)

    if workbook_path.resolve() != output_path.resolve():
        shutil.copy2(workbook_path, output_path)

    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    meter_row_map = find_meter_rows(ws)
    amount_row, amount_col = find_bill_amount_target(ws)

    matched = 0
    missing_in_bill: List[str] = []
    extra_in_bill = sorted(set(bill_rows) - set(meter_row_map))
    multiplier_changes: List[Dict[str, object]] = []
    usage_mismatches: List[Dict[str, object]] = []
    issues: List[ValidationIssue] = []

    workbook_usage_from_inputs = 0.0

    for meter, row_idx in meter_row_map.items():
        if meter not in bill_rows:
            missing_in_bill.append(meter)
            continue

        bill_row = bill_rows[meter]
        prev_cell = ws.cell(row_idx, 3)
        curr_cell = ws.cell(row_idx, 4)
        mult_cell = ws.cell(row_idx, 6)

        old_multiplier = safe_float(mult_cell.value)

        prev_cell.value = bill_row.prev_read
        curr_cell.value = bill_row.curr_read
        mult_cell.value = bill_row.multiplier

        calculated_usage = (bill_row.curr_read - bill_row.prev_read) * bill_row.multiplier
        workbook_usage_from_inputs += calculated_usage

        if old_multiplier is not None and abs(old_multiplier - bill_row.multiplier) > 1e-9:
            multiplier_changes.append(
                {
                    "meter": meter,
                    "workbook_multiplier": old_multiplier,
                    "bill_multiplier": bill_row.multiplier,
                    "row": row_idx,
                }
            )

        if abs(calculated_usage - bill_row.bill_usage_lbs) > 0.11:
            usage_mismatches.append(
                {
                    "meter": meter,
                    "row": row_idx,
                    "usage_from_inputs": calculated_usage,
                    "usage_from_bill": bill_row.bill_usage_lbs,
                }
            )
            issues.append(
                ValidationIssue(
                    level="warning",
                    message=(
                        f"Meter {meter} row {row_idx}: usage from inputs "
                        f"({calculated_usage:,.3f}) differs from bill usage "
                        f"({bill_row.bill_usage_lbs:,.3f})."
                    ),
                )
            )

        matched += 1

    ws.cell(amount_row, amount_col).value = current_charges
    wb.save(output_path)

    if abs(workbook_usage_from_inputs - bill_total_usage_lbs) > 0.11:
        issues.append(
            ValidationIssue(
                level="warning",
                message=(
                    "Total usage from workbook inputs does not match the bill total usage: "
                    f"{workbook_usage_from_inputs:,.3f} vs {bill_total_usage_lbs:,.3f}"
                ),
            )
        )

    result = GenerationResult(
        output_workbook=str(output_path),
        report_path=None,
        matched_meter_count=matched,
        workbook_meter_count=len(meter_row_map),
        bill_meter_count=len(bill_rows),
        missing_in_bill=missing_in_bill,
        extra_in_bill=extra_in_bill,
        multiplier_changes=multiplier_changes,
        usage_mismatches=usage_mismatches,
        bill_total_usage_lbs=bill_total_usage_lbs,
        workbook_total_usage_lbs_from_inputs=workbook_usage_from_inputs,
        current_charges=current_charges,
        bill_period=bill_period,
        issues=issues,
    )
    result.report_path = generate_validation_report(report_path, result, bill_rows)
    return result


def format_result(result: GenerationResult) -> str:
    lines = [
        "Generation complete.",
        f"Output workbook: {result.output_workbook}",
        f"Matched meters: {result.matched_meter_count} / workbook {result.workbook_meter_count} / bill {result.bill_meter_count}",
        f"Bill current charges: ${result.current_charges:,.2f}",
        f"Bill total usage: {result.bill_total_usage_lbs:,.3f} LB",
        f"Workbook usage from inputs: {result.workbook_total_usage_lbs_from_inputs:,.3f} LB",
    ]
    if result.bill_period:
        lines.append(f"Bill period: {result.bill_period}")
    if result.report_path:
        lines.append(f"Validation report: {result.report_path}")
    if result.missing_in_bill:
        lines.append("Meters in workbook but not in bill: " + ", ".join(result.missing_in_bill))
    if result.extra_in_bill:
        lines.append("Meters in bill but not in workbook: " + ", ".join(result.extra_in_bill))
    if result.multiplier_changes:
        lines.append(f"Multiplier changes applied: {len(result.multiplier_changes)}")
    if result.usage_mismatches:
        lines.append(f"Usage mismatches to review: {len(result.usage_mismatches)}")
    if result.issues:
        lines.append("Warnings:")
        lines.extend(f"- {issue.message}" for issue in result.issues[:20])
    return "\n".join(lines)


class SteamApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("CSU Steam Bill Automation App")
        self.geometry("920x680")
        self.minsize(860, 620)

        self.pdf_var = tk.StringVar()
        self.workbook_var = tk.StringVar()
        self.output_var = tk.StringVar()
        self.report_var = tk.StringVar()

        self._build_ui()

    def _build_ui(self) -> None:
        outer = ttk.Frame(self, padding=16)
        outer.pack(fill="both", expand=True)

        title = ttk.Label(
            outer,
            text="CSU Steam Bill Automation App",
            font=("Segoe UI", 16, "bold"),
        )
        title.pack(anchor="w")

        subtitle = ttk.Label(
            outer,
            text="Reads a Corix steam bill PDF and fills the team's trusted Excel workbook without changing formulas.",
            wraplength=840,
        )
        subtitle.pack(anchor="w", pady=(4, 14))

        self._add_picker_row(outer, "Bill PDF", self.pdf_var, self._pick_pdf)
        self._add_picker_row(outer, "Trusted Workbook / Template", self.workbook_var, self._pick_workbook)
        self._add_picker_row(outer, "Output Workbook", self.output_var, self._pick_output)
        self._add_picker_row(outer, "Validation Report (optional JSON)", self.report_var, self._pick_report)

        rules = ttk.LabelFrame(outer, text="Locked Business Rules", padding=10)
        rules.pack(fill="x", pady=(8, 12))
        ttk.Label(
            rules,
            text=(
                "• Ignore trailing letters like E in meter readings\n"
                "• Always use the bill multiplier\n"
                "• Use Current Charges as bill amount\n"
                "• Update only workbook input cells\n"
                "• Preserve the workbook's formulas and summary logic"
            ),
            justify="left",
        ).pack(anchor="w")

        actions = ttk.Frame(outer)
        actions.pack(fill="x", pady=(0, 10))
        ttk.Button(actions, text="Generate Workbook", command=self.generate).pack(side="left")
        ttk.Button(actions, text="Open Folder", command=self.open_output_folder).pack(side="left", padx=(8, 0))
        ttk.Button(actions, text="Clear Log", command=self.clear_log).pack(side="left", padx=(8, 0))

        log_frame = ttk.LabelFrame(outer, text="Validation / Run Log", padding=8)
        log_frame.pack(fill="both", expand=True)

        self.log_box = scrolledtext.ScrolledText(log_frame, wrap="word", font=("Consolas", 10))
        self.log_box.pack(fill="both", expand=True)

        self.log("Ready. Choose a PDF and a trusted workbook to begin.")

    def _add_picker_row(self, parent, label_text: str, variable: tk.StringVar, command) -> None:
        row = ttk.Frame(parent)
        row.pack(fill="x", pady=4)

        ttk.Label(row, text=label_text, width=28).pack(side="left")
        entry = ttk.Entry(row, textvariable=variable)
        entry.pack(side="left", fill="x", expand=True, padx=(0, 8))
        ttk.Button(row, text="Browse", command=command).pack(side="left")

    def _pick_pdf(self) -> None:
        path = filedialog.askopenfilename(
            title="Select steam bill PDF",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
        )
        if path:
            self.pdf_var.set(path)
            if not self.output_var.get():
                self.output_var.set(self._default_output_path(path))

    def _pick_workbook(self) -> None:
        path = filedialog.askopenfilename(
            title="Select trusted workbook/template",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.workbook_var.set(path)

    def _pick_output(self) -> None:
        path = filedialog.asksaveasfilename(
            title="Save output workbook as",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.output_var.set(path)

    def _pick_report(self) -> None:
        path = filedialog.asksaveasfilename(
            title="Save validation report as",
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
        )
        if path:
            self.report_var.set(path)

    def _default_output_path(self, pdf_path: str) -> str:
        pdf = Path(pdf_path)
        output_name = pdf.stem.replace(" Paid", "").replace(" ", "_") + "_Generated.xlsx"
        return str(pdf.with_name(output_name))

    def log(self, text: str) -> None:
        self.log_box.insert("end", text + "\n")
        self.log_box.see("end")
        self.update_idletasks()

    def clear_log(self) -> None:
        self.log_box.delete("1.0", "end")

    def open_output_folder(self) -> None:
        target = self.output_var.get().strip()
        if not target:
            messagebox.showinfo("Open Folder", "Choose an output file first.")
            return
        folder = str(Path(target).resolve().parent)
        if sys.platform.startswith("win"):
            import os
            os.startfile(folder)
        elif sys.platform == "darwin":
            import subprocess
            subprocess.Popen(["open", folder])
        else:
            import subprocess
            subprocess.Popen(["xdg-open", folder])

    def generate(self) -> None:
        pdf_path = Path(self.pdf_var.get().strip())
        workbook_path = Path(self.workbook_var.get().strip())
        output_path = Path(self.output_var.get().strip())
        report_path = Path(self.report_var.get().strip()) if self.report_var.get().strip() else None

        if not pdf_path.exists():
            messagebox.showerror("Missing PDF", "Please choose a valid bill PDF.")
            return
        if not workbook_path.exists():
            messagebox.showerror("Missing Workbook", "Please choose a valid trusted workbook/template.")
            return
        if not output_path.parent.exists():
            messagebox.showerror("Invalid Output", "The output folder does not exist.")
            return

        self.clear_log()
        self.log("Starting generation...")
        self.log(f"PDF: {pdf_path}")
        self.log(f"Workbook: {workbook_path}")
        self.log(f"Output: {output_path}")
        if report_path:
            self.log(f"Report: {report_path}")

        try:
            result = update_workbook_from_bill(pdf_path, workbook_path, output_path, report_path)
            self.log("")
            self.log(format_result(result))
            self.log("")
            self.log("Done.")
            messagebox.showinfo("Success", "Workbook generated successfully.")
        except Exception as exc:
            self.log("")
            self.log("ERROR")
            self.log(str(exc))
            self.log("")
            self.log(traceback.format_exc())
            messagebox.showerror("Generation failed", str(exc))


def main() -> None:
    app = SteamApp()
    app.mainloop()


if __name__ == "__main__":
    main()